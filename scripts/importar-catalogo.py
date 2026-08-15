#!/usr/bin/env python3
"""
Importa la hoja PRODUCTOS de la planilla de Excel/Google Sheets
y genera src/data/productos.seed.json, que la app usa para sembrar
la base de datos local la primera vez que se abre.

Uso:
    pip install openpyxl
    python3 scripts/importar-catalogo.py "JULIO 2026 nueva ECDM - CON BASE DE DATOS.xlsx"

Columnas esperadas en la hoja PRODUCTOS:
    A CODIGOS | B PRODUCTOS | C PROVEEDOR | D FECHA (compra)
    E PRECIO COMPRA | F RENTABILIDAD | G PRECIO VENTA | H FECHA (actualizacion precio)
"""
import json
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instalalo con: pip install openpyxl")

RAIZ = Path(__file__).resolve().parent.parent
SALIDA = RAIZ / "public" / "productos.seed.json"


def texto(valor):
    if valor is None:
        return ""
    return " ".join(str(valor).split()).strip()


def numero(valor):
    """Convierte a float. Tolera '$ 1.234,50', '1,234.50' y texto suelto."""
    if isinstance(valor, (int, float)):
        return round(float(valor), 2)
    limpio = texto(valor)
    if not limpio:
        return None
    limpio = re.sub(r"[^\d,.\-]", "", limpio)
    if not limpio:
        return None
    # Formato argentino: el punto separa miles y la coma los decimales.
    if "," in limpio and "." in limpio:
        limpio = limpio.replace(".", "").replace(",", ".")
    elif "," in limpio:
        limpio = limpio.replace(",", ".")
    try:
        return round(float(limpio), 2)
    except ValueError:
        return None


def fecha(valor):
    """Devuelve ISO yyyy-mm-dd. La planilla mezcla fechas reales y texto."""
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    limpio = texto(valor).replace("//", "/").replace(" ", "")
    if not limpio:
        return None
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", limpio)
    if not m:
        return None
    dia, mes, anio = (int(g) for g in m.groups())
    if anio < 100:
        anio += 2000
    try:
        return datetime(anio, mes, dia).date().isoformat()
    except ValueError:
        return None


def normalizar(cadena):
    """Minusculas sin acentos, para buscar sin que moleste la tilde."""
    sin_tildes = unicodedata.normalize("NFD", cadena)
    sin_tildes = "".join(c for c in sin_tildes if unicodedata.category(c) != "Mn")
    return sin_tildes.lower()


def main():
    if len(sys.argv) < 2:
        sys.exit(f"Uso: python3 {sys.argv[0]} <planilla.xlsx>")
    ruta = Path(sys.argv[1])
    if not ruta.exists():
        sys.exit(f"No encuentro el archivo: {ruta}")

    libro = openpyxl.load_workbook(ruta, data_only=True)
    if "PRODUCTOS" not in libro.sheetnames:
        sys.exit("La planilla no tiene una hoja llamada PRODUCTOS")
    hoja = libro["PRODUCTOS"]

    productos = []
    vistos = {}
    duplicados = 0

    for fila in range(2, hoja.max_row + 1):
        codigo = texto(hoja.cell(fila, 1).value).upper()
        if not codigo:
            continue
        descripcion = texto(hoja.cell(fila, 2).value)
        if not descripcion:
            continue

        registro = {
            "codigo": codigo,
            "descripcion": descripcion,
            "proveedor": texto(hoja.cell(fila, 3).value) or None,
            "fechaCompra": fecha(hoja.cell(fila, 4).value),
            "precioCompra": numero(hoja.cell(fila, 5).value),
            "rentabilidad": numero(hoja.cell(fila, 6).value),
            "precioVenta": numero(hoja.cell(fila, 7).value),
            "fechaPrecioVenta": fecha(hoja.cell(fila, 8).value),
            "busqueda": normalizar(f"{codigo} {descripcion}"),
            "stock": None,
            "activo": True,
        }

        if codigo in vistos:
            duplicados += 1
            # Nos quedamos con la fila mas nueva (la que tenga fecha de precio mayor).
            anterior = productos[vistos[codigo]]
            nueva = registro["fechaPrecioVenta"] or ""
            vieja = anterior["fechaPrecioVenta"] or ""
            if nueva >= vieja:
                productos[vistos[codigo]] = registro
            continue

        vistos[codigo] = len(productos)
        productos.append(registro)

    productos.sort(key=lambda p: p["descripcion"])

    SALIDA.parent.mkdir(parents=True, exist_ok=True)
    SALIDA.write_text(
        json.dumps(productos, ensure_ascii=False, indent=0, separators=(",", ":")),
        encoding="utf-8",
    )

    sin_compra = sum(1 for p in productos if not p["precioCompra"])
    sin_venta = sum(1 for p in productos if not p["precioVenta"])
    print(f"Productos exportados : {len(productos)}")
    print(f"Codigos duplicados   : {duplicados} (se conservo el mas reciente)")
    print(f"Sin precio de compra : {sin_compra}")
    print(f"Sin precio de venta  : {sin_venta}")
    print(f"Archivo              : {SALIDA.relative_to(RAIZ)}")


if __name__ == "__main__":
    main()
