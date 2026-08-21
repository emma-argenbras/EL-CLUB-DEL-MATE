#!/usr/bin/env python3
"""
Importa las ventas, egresos y gastos de un mes ya cargado en la planilla
vieja (una hoja por turno: "1M", "1T", "2M", "2T", ... + hoja TOTALES)
y genera un seed historico para que la app lo cargue solo, una vez.

Tambien genera el conteo de billetes de apertura y cierre de cada turno.

Uso:
    pip install openpyxl
    python3 scripts/importar-historico.py "planilla.xlsx" 2026 8

Genera:
    public/historico-2026-08.seed.json
    public/arqueos-2026-08.seed.json

COMO CARGA LA PLANILLA CADA VENTA
--------------------------------
La hoja de cada turno tiene, por fila: codigo (B), descripcion (C),
cantidad (D), precio unitario (E), TOTAL $ (F) y TARJETAS (G). Segun
como se cobro, la fila se completa de formas distintas, y hay que
reconocerlas todas o se pierden ventas:

  1. Efectivo             F = total en pesos, G vacio.
  2. Tarjeta / transferencia
                          F = el medio escrito a mano ("DEBITO", "QR"),
                          G = total cobrado.
  3. Pago partido         F = parte en efectivo, G = parte con tarjeta
                          (F + G = cantidad x precio).
  4. Varios productos en un solo pago con tarjeta
                          El medio queda escrito repartido en varias
                          filas ("CREDITO" / "TRES" / "CTAS") y el total
                          cobrado aparece una sola vez en G, en alguna
                          fila del grupo. Las demas filas del grupo son
                          los productos, sin total propio. Si el total
                          cobrado no coincide con la suma de los precios
                          de lista, la diferencia es un recargo por
                          financiacion (o un descuento) y se anota como
                          una linea aparte, para que la plata cierre sin
                          ensuciar el precio de cada producto.

Ademas la planilla a veces anota la misma venta en F y en G (queda
contada dos veces). Se detecta y se cuenta una sola vez.
"""
import json
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instalalo con: pip install openpyxl")

RAIZ = Path(__file__).resolve().parent.parent

# Los textos con que el vendedor anotaba el medio de pago en la columna F
# de cada fila de venta (a mano, con variantes y errores de tipeo).
MEDIOS = {
    "": "EFECTIVO",
    "TRANSFERENCIA": "TRANSFERENCIA",
    "TRNASFERENCIA": "TRANSFERENCIA",
    "DEBITO": "DEBITO",
    "DEBTIO": "DEBITO",
    "QR": "QR",
    "CREDITO": "CREDITO",
    "CRED": "CREDITO",
    "CRED UN PAGO": "CREDITO",
    "UN PAGO": "CREDITO",
    "CONSUMAX": "CREDITO",
}

# Para reconocer el medio cuando quedo escrito repartido en varias filas
# ("CREDITO" / "TRES" / "CTAS") se busca la palabra dentro del texto junto.
PALABRAS_MEDIO = [
    ("TRANSFERENCIA", "TRANSFERENCIA"),
    ("TRNASFERENCIA", "TRANSFERENCIA"),
    ("TRANSFRENCIA", "TRANSFERENCIA"),
    ("TRANSF", "TRANSFERENCIA"),
    ("CONSUMAX", "CREDITO"),
    ("CREDITO", "CREDITO"),
    ("CRED", "CREDITO"),
    ("DEBITO", "DEBITO"),
    ("DEBTIO", "DEBITO"),
    ("DEB", "DEBITO"),
    ("QR", "QR"),
]

# Categorias de gasto que se detectan por palabras clave en el concepto.
CATEGORIA_POR_PALABRA = {
    "ALQUILER": "ALQUILER",
    "LUZ": "SERVICIOS",
    "INTERNET": "SERVICIOS",
    "CONTADOR": "CONTADOR",
    "IMPUESTO": "IMPUESTOS",
    "AFIP": "IMPUESTOS",
    "SUELDO": "SUELDOS",
    "GABI": "SUELDOS",
    "ELENA": "SUELDOS",
}

FIJOS = {"ALQUILER", "SERVICIOS", "SUELDOS", "CONTADOR", "IMPUESTOS", "MANTENIMIENTO"}

# Primera fila de venta y ultima fila hasta donde se busca. La planilla
# deja 20 renglones libres, pero en un dia movido se sigue escribiendo
# mas abajo: hay que leer hasta donde realmente haya algo cargado.
FILA_PRIMERA_VENTA = 8
FILA_ULTIMA_BUSQUEDA = 60

# Codigo con que se anota la diferencia entre lo cobrado y la suma de los
# precios de lista cuando se pagan varios productos juntos con tarjeta.
CODIGO_AJUSTE = "AJUSTE"


def num(valor):
    if isinstance(valor, bool):
        return None
    return valor if isinstance(valor, (int, float)) else None


def txt(valor):
    return " ".join(str(valor).split()).strip() if valor is not None else ""


def normalizar(cadena):
    sin_tildes = unicodedata.normalize("NFD", cadena)
    sin_tildes = "".join(c for c in sin_tildes if unicodedata.category(c) != "Mn")
    return sin_tildes.lower()


def categoria_de(concepto):
    mayus = concepto.upper()
    for palabra, cat in CATEGORIA_POR_PALABRA.items():
        if palabra in mayus:
            return cat
    return "PROVEEDORES"


def medio_de(texto, por_tarjeta=False):
    """Reconoce el medio de pago, aunque este escrito repartido o con
    errores de tipeo. Si el importe estaba en la columna de TARJETAS, un
    texto que no reconocemos no puede ser efectivo: queda transferencia,
    que es lo mas usado en el local."""
    limpio = txt(texto).upper()
    if limpio in MEDIOS and not (por_tarjeta and limpio == ""):
        return MEDIOS[limpio]
    for palabra, medio in PALABRAS_MEDIO:
        if palabra in limpio:
            return medio
    return "TRANSFERENCIA" if por_tarjeta else "EFECTIVO"


def leer_filas(ws):
    """Devuelve las filas de venta con contenido, ya interpretadas."""
    filas = []
    for r in range(FILA_PRIMERA_VENTA, FILA_ULTIMA_BUSQUEDA + 1):
        codigo = txt(ws.cell(r, 2).value).upper()
        descripcion = txt(ws.cell(r, 3).value)
        cantidad = num(ws.cell(r, 4).value)
        precio = num(ws.cell(r, 5).value)
        f_val = ws.cell(r, 6).value
        f_num = num(f_val)
        f_txt = txt(f_val).upper() if f_num is None else ""
        g_num = num(ws.cell(r, 7).value)

        if not (codigo or descripcion or cantidad or precio or f_num or f_txt or g_num):
            continue

        filas.append(
            {
                "fila": r,
                "codigo": codigo,
                "descripcion": descripcion,
                "cantidad": cantidad,
                "precio": precio,
                "f_num": f_num,
                "f_txt": f_txt,
                "g_num": g_num,
            }
        )
    return filas


def lista_de(fila):
    """Precio de lista de la fila: cantidad x precio unitario."""
    return (fila["cantidad"] or 1) * (fila["precio"] or 0)


def venta_de(fila, total, medio, fecha, turno, costos, sin_costo):
    cantidad = fila["cantidad"] or 1
    if cantidad <= 0:
        cantidad = 1
    precio = fila["precio"]
    if not precio:
        precio = round(total / cantidad, 2)

    codigo = fila["codigo"]
    descripcion = fila["descripcion"] or codigo or "Venta sin descripción"
    costo = costos.get(codigo) if codigo else None
    if codigo and costo is None:
        entrada = sin_costo.setdefault(
            codigo, {"descripcion": descripcion, "unidades": 0, "monto": 0}
        )
        entrada["unidades"] += cantidad
        entrada["monto"] += total

    return {
        "fecha": fecha,
        "turno": turno,
        "hora": "",
        "codigo": codigo,
        "descripcion": descripcion,
        "cantidad": cantidad,
        "precioUnitario": precio,
        "costoUnitario": costo,
        "medioPago": medio,
        "total": round(total, 2),
        "vendedor": None,
    }


def ajuste_de(monto, medio, fecha, turno, detalle=None):
    """Linea que absorbe el recargo por financiacion (o el descuento) de
    un pago agrupado, para que la plata cierre sin tocar el precio de
    lista de cada producto. Sin costo de mercaderia: no es un producto."""
    positivo = monto > 0
    return {
        "fecha": fecha,
        "turno": turno,
        "hora": "",
        "codigo": CODIGO_AJUSTE,
        "descripcion": detalle
        or ("Recargo por financiación" if positivo else "Descuento en la venta"),
        "cantidad": 1,
        "precioUnitario": round(monto, 2),
        "costoUnitario": 0,
        "medioPago": medio,
        "total": round(monto, 2),
        "vendedor": None,
    }


def es_cobro_completo(fila):
    """Una fila en efectivo esta 'completa' cuando lo anotado en F es todo
    lo que valia el producto. Si es bastante menos, fue una entrega a
    cuenta y el resto se cobro junto con las filas de al lado."""
    lista = lista_de(fila)
    if lista <= 0:
        return True  # sin precio de lista no hay con que comparar
    return fila["f_num"] >= lista * 0.9


def procesar_bloque(bloque, fecha, turno, costos, sin_costo, avisos):
    """Un bloque es un conjunto de filas seguidas que se cobraron juntas.
    Puede tener filas sin total propio (los productos) y una fila con el
    total realmente cobrado."""
    ventas = []

    # Filas cuyo importe en TARJETAS coincide con su propio precio de
    # lista: son ventas sueltas, no arrastran a las de al lado.
    sueltas, totales_grupo = [], []
    for idx, f in enumerate(bloque):
        if f["g_num"] is None:
            continue
        if abs(f["g_num"] - lista_de(f)) < 1:
            sueltas.append(idx)
        else:
            totales_grupo.append(idx)

    asignado = set(sueltas)
    for idx in sueltas:
        f = bloque[idx]
        ventas.append(
            venta_de(f, f["g_num"], medio_de(f["f_txt"], True), fecha, turno, costos, sin_costo)
        )

    # Cada fila que quedo sin total propio se suma al grupo mas cercano.
    grupos = {idx: [idx] for idx in totales_grupo}
    for idx in range(len(bloque)):
        if idx in asignado or idx in grupos:
            continue
        if not totales_grupo:
            continue
        cercano = min(totales_grupo, key=lambda t: abs(t - idx))
        grupos[cercano].append(idx)
        asignado.add(idx)

    for idx_total, indices in grupos.items():
        indices = sorted(indices)
        filas_grupo = [bloque[k] for k in indices]
        cobrado_tarjeta = sum(f["g_num"] or 0 for f in filas_grupo)
        cobrado_efectivo = sum(f["f_num"] or 0 for f in filas_grupo)
        suma_lista = sum(lista_de(f) for f in filas_grupo)
        medio = medio_de(" ".join(f["f_txt"] for f in filas_grupo), True)

        for f in filas_grupo:
            ventas.append(venta_de(f, lista_de(f), medio, fecha, turno, costos, sin_costo))

        # Parte entregada en efectivo dentro de un pago con tarjeta: se
        # mueve del medio de pago de la tarjeta al de efectivo, para que
        # el cierre de caja del turno cierre igual que en la planilla.
        if cobrado_efectivo > 0:
            ventas.append(ajuste_de(cobrado_efectivo, "EFECTIVO", fecha, turno, "Parte pagada en efectivo"))
            ventas.append(ajuste_de(-cobrado_efectivo, medio, fecha, turno, "Parte pagada en efectivo"))

        diferencia = round(cobrado_tarjeta + cobrado_efectivo - suma_lista, 2)
        if abs(diferencia) >= 1:
            ventas.append(ajuste_de(diferencia, medio, fecha, turno))
            signo = "recargo" if diferencia > 0 else "descuento"
            avisos.append(
                f"{fecha} {turno} filas {filas_grupo[0]['fila']}-{filas_grupo[-1]['fila']}: "
                f"{len(indices)} producto(s) por ${suma_lista:,.0f} cobrados "
                f"${cobrado_tarjeta + cobrado_efectivo:,.0f} ({signo} de ${abs(diferencia):,.0f})".replace(",", ".")
            )

    # Filas anotadas pero que la planilla no suma en ningun subtotal: son
    # renglones empezados y abandonados. Se dejan afuera (si no, la caja
    # del turno daria de mas) pero se avisan para poder revisarlos.
    for idx, f in enumerate(bloque):
        if idx in asignado or idx in grupos:
            continue
        avisos.append(
            f"{fecha} {turno} fila {f['fila']}: {f['codigo'] or f['descripcion'][:20]} quedo sin "
            f"importe y la planilla no lo suma; NO se importa."
        )

    return ventas


def procesar_ventas(filas, fecha, turno, costos, sin_costo, avisos):
    """Recorre las filas de un turno y arma las ventas, reconociendo los
    cuatro modos de carga descriptos arriba."""
    ventas = []
    i = 0
    while i < len(filas):
        fila = filas[i]

        # --- Pago partido, o la misma venta anotada dos veces ---
        if fila["f_num"] is not None and fila["g_num"] is not None:
            esperado = lista_de(fila)
            if esperado and abs(fila["f_num"] + fila["g_num"] - esperado) < 1:
                # Parte en efectivo y parte con tarjeta. La app guarda un
                # solo medio de pago por venta, asi que se parte en dos
                # renglones: el producto (con su costo) por la parte en
                # efectivo, y el resto como una linea sin unidades ni
                # costo, para que ni la caja ni el margen se desajusten.
                efectivo = venta_de(
                    fila, fila["f_num"], "EFECTIVO", fecha, turno, costos, sin_costo
                )
                efectivo["precioUnitario"] = round(fila["f_num"] / efectivo["cantidad"], 2)
                ventas.append(efectivo)

                # El {} descarta el aviso de "sin costo": ya lo conto la
                # mitad en efectivo, este renglon es la misma venta.
                resto = venta_de(fila, fila["g_num"], "TRANSFERENCIA", fecha, turno, costos, {})
                resto["cantidad"] = 0
                resto["precioUnitario"] = round(fila["g_num"], 2)
                resto["costoUnitario"] = 0
                resto["descripcion"] = f"{efectivo['descripcion']} (parte con tarjeta)"
                ventas.append(resto)
            else:
                # La planilla la anoto en las dos columnas: se cuenta una vez.
                avisos.append(
                    f"{fecha} {turno} fila {fila['fila']}: {fila['codigo']} figura en efectivo "
                    f"(${fila['f_num']:,.0f}) y en tarjetas (${fila['g_num']:,.0f}); se cuenta una sola vez."
                    .replace(",", ".")
                )
                ventas.append(
                    venta_de(fila, fila["f_num"], "EFECTIVO", fecha, turno, costos, sin_costo)
                )
            i += 1
            continue

        # --- Efectivo cobrado entero en su propia fila ---
        if fila["f_num"] is not None and es_cobro_completo(fila):
            if fila["f_num"] > 0:
                ventas.append(
                    venta_de(fila, fila["f_num"], "EFECTIVO", fecha, turno, costos, sin_costo)
                )
            i += 1
            continue

        # --- Bloque cobrado junto ---
        # Arranca en una fila con tarjeta (o en una entrega a cuenta) y se
        # extiende mientras las filas siguientes no sean cobros completos
        # en efectivo ni pagos partidos, que empiezan otra operacion.
        inicio = i
        i += 1
        while i < len(filas):
            siguiente = filas[i]
            if siguiente["f_num"] is not None and siguiente["g_num"] is not None:
                break
            if siguiente["f_num"] is not None and es_cobro_completo(siguiente):
                break
            i += 1

        ventas.extend(
            procesar_bloque(filas[inicio:i], fecha, turno, costos, sin_costo, avisos)
        )

    return ventas


def leer_arqueo(ws, col_denominacion, col_monto):
    """El conteo de billetes: la planilla anota el MONTO por denominacion
    (no la cantidad), mas una fila GRANDE con los billetes altos juntos y
    otra de MONEDA. La app guarda cantidades, asi que se divide."""
    billetes = {}
    monedas = 0
    grande = 0
    for r in range(7, 17):
        etiqueta = txt(ws.cell(r, col_denominacion).value).upper()
        monto = num(ws.cell(r, col_monto).value)
        if monto is None or monto == 0:
            continue
        if etiqueta == "GRANDE":
            grande = monto
        elif etiqueta == "MONEDA":
            monedas += monto
        else:
            denominacion = num(ws.cell(r, col_denominacion).value)
            if denominacion and denominacion > 0:
                billetes[str(int(denominacion))] = int(monto // denominacion)
                monedas += monto % denominacion

    # "GRANDE" son los billetes de mayor valor sin discriminar: se cargan
    # como billetes de 2000 (el mas alto de uso corriente en la planilla)
    # y el resto que no divide justo va a monedas, para no perder plata.
    if grande:
        billetes["2000"] = billetes.get("2000", 0) + int(grande // 2000)
        monedas += grande % 2000

    if not billetes and not monedas:
        return None
    return {"billetes": billetes, "monedas": round(monedas, 2)}


def main():
    if len(sys.argv) < 4:
        sys.exit(f"Uso: python3 {sys.argv[0]} <planilla.xlsx> <anio> <mes>")
    ruta = Path(sys.argv[1])
    anio = int(sys.argv[2])
    mes = int(sys.argv[3])
    if not ruta.exists():
        sys.exit(f"No encuentro el archivo: {ruta}")

    libro = openpyxl.load_workbook(ruta, data_only=True)

    # Precios de compra actuales, para asignar el costo de cada venta.
    costos = {}
    if "PRODUCTOS" in libro.sheetnames:
        hp = libro["PRODUCTOS"]
        for r in range(2, hp.max_row + 1):
            codigo = txt(hp.cell(r, 1).value).upper()
            compra = num(hp.cell(r, 5).value)
            if codigo:
                costos[codigo] = compra

    jornadas = []
    ventas = []
    movimientos = []
    arqueos = []
    sin_costo = {}
    avisos = []
    control = []  # (hoja, subtotal planilla efectivo/tarjeta, lo que importamos)

    for dia in range(1, 32):
        for turno in ("M", "T"):
            nombre_hoja = f"{dia}{turno}"
            if nombre_hoja not in libro.sheetnames:
                continue
            ws = libro[nombre_hoja]

            try:
                fecha = datetime(anio, mes, dia).date().isoformat()
            except ValueError:
                continue  # dia invalido para el mes (ej. 31 de un mes de 30)

            caja_inicial = num(ws.cell(6, 6).value) or 0
            filas = leer_filas(ws)
            ventas_dia = procesar_ventas(filas, fecha, turno, costos, sin_costo, avisos)

            # Egresos chicos pagados con la caja del turno (H8:H17, detalle en I).
            egresos_dia = []
            for r in range(8, 18):
                monto = num(ws.cell(r, 8).value)
                if monto and monto > 0:
                    concepto = txt(ws.cell(r, 9).value) or "Egreso"
                    egresos_dia.append({"concepto": concepto, "monto": monto})

            # Efectivo que se pasa de la caja del turno a la caja grande (H21:H27).
            pases_dia = []
            for r in range(21, 28):
                monto = num(ws.cell(r, 8).value)
                if monto and monto > 0:
                    concepto = txt(ws.cell(r, 9).value) or "Pase a caja grande"
                    pases_dia.append({"concepto": concepto, "monto": monto})

            if not ventas_dia and not egresos_dia and not pases_dia and caja_inicial == 0:
                continue  # turno que nunca se abrio

            # Control contra los subtotales que calcula la propia planilla.
            sub_ef = num(ws.cell(3, 6).value) or 0
            sub_tj = num(ws.cell(3, 7).value) or 0
            imp_ef = sum(v["total"] for v in ventas_dia if v["medioPago"] == "EFECTIVO")
            imp_tj = sum(v["total"] for v in ventas_dia if v["medioPago"] != "EFECTIVO")
            control.append((nombre_hoja, sub_ef, imp_ef, sub_tj, imp_tj))

            arqueo_apertura = leer_arqueo(ws, 11, 12)  # K / L
            arqueo_cierre = leer_arqueo(ws, 14, 15)  # N / O

            indice_jornada = len(jornadas)
            jornadas.append(
                {
                    "fecha": fecha,
                    "turno": turno,
                    "estado": "cerrado",
                    "vendedor": None,
                    "cajaInicial": caja_inicial,
                    "horaApertura": None,
                    "horaCierre": None,
                    "arqueoApertura": arqueo_apertura
                    or {"billetes": {}, "monedas": caja_inicial},
                    "arqueoCierre": arqueo_cierre,
                    "notas": "Importado desde la planilla de Google Sheets.",
                }
            )
            arqueos.append(
                {
                    "fecha": fecha,
                    "turno": turno,
                    "arqueoApertura": arqueo_apertura,
                    "arqueoCierre": arqueo_cierre,
                }
            )

            for v in ventas_dia:
                v["_jornadaIndice"] = indice_jornada
                ventas.append(v)

            for e in egresos_dia:
                # Que sea fijo o variable depende de QUE se pago, no de con
                # que caja se pago: un sueldo sacado de la caja del turno
                # sigue siendo un gasto fijo. Si se cuenta como variable, el
                # margen de contribucion sale mas bajo de lo real.
                categoria = categoria_de(e["concepto"])
                movimientos.append(
                    {
                        "fecha": fecha,
                        "tipo": "EGRESO_CAJA",
                        "concepto": e["concepto"],
                        "monto": e["monto"],
                        "categoria": categoria,
                        "_jornadaIndice": indice_jornada,
                        "esVariable": categoria not in FIJOS,
                    }
                )
            for p in pases_dia:
                movimientos.append(
                    {
                        "fecha": fecha,
                        "tipo": "A_CAJA_GRANDE",
                        "concepto": p["concepto"],
                        "monto": p["monto"],
                        "categoria": None,
                        "_jornadaIndice": indice_jornada,
                        "esVariable": False,
                    }
                )

    # Gastos pagados desde la caja grande (hoja TOTALES, columnas J concepto / K monto / L fecha).
    if "TOTALES" in libro.sheetnames:
        wt = libro["TOTALES"]
        for r in range(6, 200):
            concepto = txt(wt.cell(r, 10).value)
            monto = num(wt.cell(r, 11).value)
            fecha_celda = wt.cell(r, 12).value
            if not concepto or not monto:
                continue
            fecha = (
                fecha_celda.date().isoformat()
                if isinstance(fecha_celda, datetime)
                else f"{anio:04d}-{mes:02d}-01"
            )
            categoria = categoria_de(concepto)
            movimientos.append(
                {
                    "fecha": fecha,
                    "tipo": "GASTO_CAJA_GRANDE",
                    "concepto": concepto,
                    "monto": monto,
                    "categoria": categoria,
                    "_jornadaIndice": None,
                    "esVariable": categoria not in FIJOS,
                }
            )

    salida = {
        "mes": f"{anio:04d}-{mes:02d}",
        "jornadas": jornadas,
        "ventas": ventas,
        "movimientos": movimientos,
    }

    destino = RAIZ / "public" / f"historico-{anio:04d}-{mes:02d}.seed.json"
    destino.write_text(
        json.dumps(salida, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    destino_arqueos = RAIZ / "public" / f"arqueos-{anio:04d}-{mes:02d}.seed.json"
    destino_arqueos.write_text(
        json.dumps(arqueos, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )

    total_ventas = sum(v["total"] for v in ventas)
    print(f"Jornadas importadas  : {len(jornadas)}")
    print(f"Ventas importadas    : {len(ventas)} por un total de ${total_ventas:,.0f}".replace(",", "."))
    print(f"Movimientos          : {len(movimientos)}")
    print(f"Archivos generados   : {destino.relative_to(RAIZ)}")
    print(f"                       {destino_arqueos.relative_to(RAIZ)}")

    # ---- Control: lo importado tiene que dar igual que la planilla ----
    print()
    print("CONTROL contra los subtotales de la propia planilla:")
    descuadres = 0
    tot_ef_p = tot_ef_i = tot_tj_p = tot_tj_i = 0
    for hoja, sub_ef, imp_ef, sub_tj, imp_tj in control:
        tot_ef_p += sub_ef
        tot_ef_i += imp_ef
        tot_tj_p += sub_tj
        tot_tj_i += imp_tj
        if abs(sub_ef - imp_ef) >= 1 or abs(sub_tj - imp_tj) >= 1:
            descuadres += 1
            print(
                f"  {hoja:5s} efectivo planilla ${sub_ef:,.0f} vs importado ${imp_ef:,.0f} | "
                f"tarjetas planilla ${sub_tj:,.0f} vs importado ${imp_tj:,.0f}".replace(",", ".")
            )
    print(f"  Efectivo  planilla ${tot_ef_p:,.0f}  importado ${tot_ef_i:,.0f}".replace(",", "."))
    print(f"  Tarjetas  planilla ${tot_tj_p:,.0f}  importado ${tot_tj_i:,.0f}".replace(",", "."))
    print(f"  Turnos con diferencia: {descuadres} de {len(control)}")

    if avisos:
        print()
        print(f"Filas que hubo que interpretar ({len(avisos)}):")
        for a in avisos:
            print(f"  {a}")

    print()
    print(f"Productos vendidos SIN costo de compra cargado ({len(sin_costo)}):")
    for codigo, d in sorted(sin_costo.items(), key=lambda kv: -kv[1]["monto"])[:20]:
        print(
            f"  {codigo:12s} {d['descripcion']:45s} {d['unidades']:>4.0f} un.  ${d['monto']:>10,.0f}".replace(",", ".")
        )


if __name__ == "__main__":
    main()
