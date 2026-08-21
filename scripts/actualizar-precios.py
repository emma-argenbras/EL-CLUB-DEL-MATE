#!/usr/bin/env python3
"""
Trae los costos y precios de la planilla "BASE DE DATOS ECDM 2026" (la
fuente de verdad de la lista de precios) al catalogo de la app.

Genera dos archivos:

  public/productos.seed.json      catalogo completo, para una instalacion nueva
  public/precios-<mes>.seed.json  parche para las apps que ya estan en uso

El parche guarda, por producto, el valor ANTERIOR y el NUEVO de cada
campo. La app aplica el nuevo solamente si lo que tiene guardado sigue
siendo igual al anterior; si alguien lo edito a mano desde la app, ese
producto se deja como esta. Asi actualizar la lista nunca pisa el
trabajo de nadie.

Uso:
    pip install openpyxl
    python3 scripts/actualizar-precios.py "BASE DE DATOS ECDM 2026.xlsx" 2026-08

COLUMNAS DE LA HOJA "PRODUCTOS" DE LA BASE
------------------------------------------
  A CODIGOS   B PRODUCTOS   C PROVEEDOR   E STOCK (=F-G)   F ENTRADAS
  G SALIDAS   H FECHA (de compra)          I PRECIO COMPRA
  J RENTABILIDAD (1,65 = 165 %)            K PRECIO VENTA
  L GANANCIA  M FECHA (de actualizacion del precio de venta)

La columna K tiene, en los productos que siguen la regla general, la
formula  =ROUNDUP((I*J+I),-2)  : el costo mas el markup, redondeado
SIEMPRE para arriba al siguiente multiplo de 100, para que quede un
numero facil de cobrar. En muchos productos el precio esta puesto a
mano, pisando esa formula; por eso la app la ofrece como sugerencia y
deja cambiarla, en vez de imponerla.
"""
import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instalalo con: pip install openpyxl")

RAIZ = Path(__file__).resolve().parent.parent
SEED = RAIZ / "public" / "productos.seed.json"

# Campos que se traen de la planilla. El resto (stock, archivado, notas)
# es de la app y no se toca nunca.
CAMPOS = ["precioCompra", "fechaCompra", "rentabilidad", "precioVenta", "fechaPrecioVenta"]


def num(valor):
    if isinstance(valor, bool):
        return None
    return valor if isinstance(valor, (int, float)) else None


def txt(valor):
    return " ".join(str(valor).split()).strip() if valor is not None else ""


def normalizar(cadena):
    sin_tildes = unicodedata.normalize("NFD", cadena)
    sin_tildes = "".join(c for c in sin_tildes if unicodedata.category(c) != "Mn")
    return sin_tildes.lower()


def fecha_iso(valor):
    """Las fechas vienen como fecha de verdad o tipeadas a mano ('25/1/2 025')."""
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    crudo = txt(valor).replace(" ", "")
    if not crudo:
        return None
    partes = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})$", crudo)
    if not partes:
        return None
    dia, mes, anio = int(partes.group(1)), int(partes.group(2)), int(partes.group(3))
    if anio < 100:
        anio += 2000
    try:
        return datetime(anio, mes, dia).date().isoformat()
    except ValueError:
        return None


def main():
    if len(sys.argv) < 3:
        sys.exit(f"Uso: python3 {sys.argv[0]} <base-de-datos.xlsx> <mes: 2026-08>")
    ruta = Path(sys.argv[1])
    mes = sys.argv[2]
    if not ruta.exists():
        sys.exit(f"No encuentro el archivo: {ruta}")

    libro = openpyxl.load_workbook(ruta, data_only=True)
    if "PRODUCTOS" not in libro.sheetnames:
        sys.exit("La planilla no tiene una hoja PRODUCTOS")
    ws = libro["PRODUCTOS"]

    filas_por_codigo = defaultdict(list)
    for r in range(2, ws.max_row + 1):
        codigo = txt(ws.cell(r, 1).value).upper()
        if not codigo:
            continue
        filas_por_codigo[codigo].append(
            {
                "codigo": codigo,
                "descripcion": txt(ws.cell(r, 2).value),
                "proveedor": txt(ws.cell(r, 3).value) or None,
                "fechaCompra": fecha_iso(ws.cell(r, 8).value),
                "precioCompra": num(ws.cell(r, 9).value),
                "rentabilidad": num(ws.cell(r, 10).value),
                "precioVenta": num(ws.cell(r, 11).value),
                "fechaPrecioVenta": fecha_iso(ws.cell(r, 13).value),
                "fila": r,
            }
        )

    # Un codigo repetido en dos productos distintos es un problema de la
    # planilla: no hay forma de saber cual gana, asi que se deja afuera
    # del parche y se avisa para corregirlo en la planilla.
    duplicados = {c: f for c, f in filas_por_codigo.items() if len(f) > 1}
    base = {c: f[0] for c, f in filas_por_codigo.items() if len(f) == 1}

    catalogo = json.load(open(SEED, encoding="utf-8"))
    por_codigo = {p["codigo"].upper(): p for p in catalogo}

    cambios = []
    nuevos = []
    resumen = defaultdict(int)

    for codigo, fila in base.items():
        actual = por_codigo.get(codigo)

        if actual is None:
            producto = {
                "codigo": codigo,
                "descripcion": fila["descripcion"] or codigo,
                "proveedor": fila["proveedor"],
                "proveedorId": None,
                "fechaCompra": fila["fechaCompra"],
                "precioCompra": fila["precioCompra"],
                "rentabilidad": fila["rentabilidad"],
                "precioVenta": fila["precioVenta"],
                "fechaPrecioVenta": fila["fechaPrecioVenta"],
                "busqueda": normalizar(f"{codigo} {fila['descripcion']}"),
                "stock": None,
                "activo": True,
            }
            nuevos.append(producto)
            catalogo.append(producto)
            continue

        anterior, nuevo = {}, {}
        for campo in CAMPOS:
            valor = fila[campo]
            if valor is None:
                continue  # la planilla no lo tiene cargado: no se borra lo que ya hay
            if valor == actual.get(campo):
                continue
            anterior[campo] = actual.get(campo)
            nuevo[campo] = valor
            resumen[campo] += 1

        if not nuevo:
            continue

        cambios.append({"codigo": codigo, "anterior": anterior, "nuevo": nuevo})
        actual.update(nuevo)

    catalogo.sort(key=lambda p: p["codigo"])
    SEED.write_text(
        json.dumps(catalogo, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )

    parche = {"mes": mes, "cambios": cambios, "nuevos": nuevos}
    destino = RAIZ / "public" / f"precios-{mes}.seed.json"
    destino.write_text(
        json.dumps(parche, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )

    print(f"Productos en la base       : {len(filas_por_codigo)}")
    print(f"Productos en el catalogo   : {len(catalogo)}")
    print(f"Productos con cambios      : {len(cambios)}")
    for campo in CAMPOS:
        if resumen[campo]:
            print(f"    {campo:20s} {resumen[campo]}")
    print(f"Productos nuevos           : {len(nuevos)}")
    for p in nuevos:
        print(f"    {p['codigo']:12s} {p['descripcion'][:40]:42s} costo {p['precioCompra']}")
    print(f"Archivos generados         : public/productos.seed.json")
    print(f"                             {destino.relative_to(RAIZ)}")

    if duplicados:
        print()
        print(f"ATENCION: {len(duplicados)} codigos repetidos en la planilla, apuntando a")
        print("productos distintos. Quedan afuera de la actualizacion hasta corregirlos,")
        print("porque no hay forma de saber cual de los dos es el bueno:")
        for codigo, filas in sorted(duplicados.items()):
            print(f"  {codigo}")
            for f in filas:
                print(
                    f"      fila {f['fila']:5d}  {f['descripcion'][:42]:44s} "
                    f"costo {f['precioCompra']}  venta {f['precioVenta']}"
                )


if __name__ == "__main__":
    main()
